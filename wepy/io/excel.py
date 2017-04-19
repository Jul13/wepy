# ==============================================================================
# title           : excel
# description     :
# author          : Amine Kamoun
# date            : 28/03/17
# version         :
# notes           :
# IDE             : PyCharm Community Edition
# ==============================================================================
import datetime
from itertools import islice

from openpyxl import load_workbook as loader_xl

from topyc.data.Feature import Feature


def read_worksheets_names(file_path):
    workbook = loader_xl(file_path)
    sheet_names = workbook.sheetnames
    # [print(s) for s in sheet_names]
    return sheet_names


def get_workbook(file_path):
    return loader_xl(file_path)


def worksheet_to_feature(wb, worksheet_name, feature_type):
    ws = wb[worksheet_name]
    nb_rows = ws.max_column
    first_date = 1

    while not isinstance(ws["A" + str(first_date)].value, datetime.datetime) and first_date <= nb_rows:
        first_date += 1

    coordinate_start = "A" + str(first_date)
    coordinate_end = ws.cell(row=ws.max_row, column=ws.max_column).coordinate
    data = get_range_values(ws, coordinate_start + ":" + coordinate_end)
    # cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = Feature(feature_type=feature_type, data=data, index=idx)
    return df


def get_range_values(ws, range_str):
    for row in ws[range_str]:
        yield tuple(c.value for c in row)
